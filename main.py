import requests
import time
import config
from openpyxl import Workbook
from func import *
from tokens import TokenList
from prelog_msg import *


tokens = TokenList(*read_token())


def get_response(url, vendor_code):
    token = tokens.get_active_token()
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Content-Length": str(len(str(vendor_code).encode())),
        "Cache-Control": "no-cache",
        "Charset": "utf-8",
        "Authorization": "Bearer " + token}
    response = requests.post(url=url + "/search-products",
                             json={"vendor_code": vendor_code},
                             headers=headers)
    try:
        return response.json()
    except ValueError:
        return response.text
    except Exception as e:
        print(e)


def main():
    counter = read_counter()
    if counter != 0:
        try:
            wb = load_workbook(config.OUTPUT_FILE_PATH)
            user_response = input(user_info + f' Continue from last saved position {counter}?'
                                  f' Press "y" (yes) or "n" (no) ').lower()
            if (user_response != 'y') and (user_response != 'yes'):
                counter = 0
                wb = Workbook()
        except KeyError as eKeyError:
            print(eKeyError)
            print(file_info + ' Excel file has been broken.'
                  'The new one will be created and started parsing from the 1st position')
            counter = 0
            wb = Workbook()
    else:
        wb = Workbook()
    ws = wb.active
    list_of_vendor_codes = [str(i[0]).replace("'", "") for i in read_input()]
    settings_list = sorted(read_settings(), reverse=True)
    position_counter = read_counter() + 1 if counter > 0 else 1
    while counter < len(list_of_vendor_codes):
        try:
            response_json = get_response(config.BASE_URL, list_of_vendor_codes[counter])
            # print(response_json)
            if type(response_json) != dict:
                if len(response_json) == 0:
                    counter += 1
                    write_counter(counter)
                    print(request_info + f" No results for the request: {list_of_vendor_codes[counter]}")
                    continue
                if "limit exceeded" in response_json.lower():
                    print(token_info + f" Limit exceeded at {tokens.get_token_order() + 1} token in your list\n" +
                          token_info + " Trying to go further with next token")
                    tokens.next_token()
                    continue
            if 'products' in response_json.keys():
                result_json = response_json["products"][0]
                str_counter = str(position_counter)
                ws["A" + str_counter] = result_json["vendor"]
                ws["B" + str_counter] = result_json["vendor_code"]
                ws["C" + str_counter] = result_json["name"]
                ws["D" + str_counter] = result_json["price"].replace('.', ',')
                if result_json["rest"] == "есть":
                    ws["E" + str_counter] = "2"
                else:
                    ws["E" + str_counter] = "0"
                wb.save(config.OUTPUT_FILE_PATH)
                print(f"{result_json['vendor_code']} success")
                position_counter += 1
            elif 'message' in response_json.keys():
                if 'unauthenticated' in response_json['message'].lower():
                    print(token_info + f' Your {tokens.get_token_order()+1} token is UNAUTHENTICATED\n' +
                          token_info + " Trying to go further with next token")
                    tokens.next_token()
                    continue
                elif 'too many attempts' in response_json['message'].lower():
                    print('Pause 5 sec')
                    time.sleep(5)
                    continue
            for setting in settings_list:
                if counter % setting[0] == 0 and counter != 0:
                    print(request_info + f" {counter + 1} positions passed. Pause {setting[1]} sec")
                    time.sleep(setting[1])
                    break
            counter += 1
            write_counter(counter)
        except Exception as e:
            print(e)
            try:
                print(request_info + f" {get_response(config.BASE_URL, list_of_vendor_codes[counter])} pause 30 minutes")
                time.sleep(1800)
            except Exception as e:
                print(e)
                print(request_info + f" Next try in {time.ctime(time.time() + 120)}")
                time.sleep(120)


if __name__ == '__main__':
    main()

